from getpass import getpass
from openpyxl import Workbook
from axl_config import AXLConfig

def get_mac_addresses(service, query):
    response = service.executeSQLQuery(sql=query)
    mac_addresses = []
    for row in response['return']['row']:
        device_name = row[0].text
        description = row[1].text
        directory_number = row[2].text
        mac_addresses.append([device_name, description, directory_number])
    return mac_addresses

def save_mac_addresses(mac_addresses, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'MAC Addresses'
    ws['A1'] = 'Device Name'
    ws['B1'] = 'Description'
    ws['C1'] = 'Directory Number'
    for mac_address in mac_addresses:
        ws.append(mac_address)
    wb.save(file_path)

if __name__ == '__main__':
    username = input('Enter Username: ')
    password = getpass("\nEnter AXL Password\n")
    directory_number_prefix = '555'

    # Instantiate AXLConfig
    axl_config = AXLConfig(username, password)

    # Create a Zeep client and service
    service = axl_config.create_service()

    query = f"SELECT d.name, d.description, n.dnorpattern as DN from device as d, numplan as n, devicenumplanmap as dnpm WHERE dnpm.fkdevice = d.pkid and dnpm.fknumplan = n.pkid and d.tkclass = 1 and n.dnorpattern LIKE '{directory_number_prefix}%'"

    mac_addresses = get_mac_addresses(service, query)
    save_mac_addresses(mac_addresses, './mac_addresses.xlsx')